import os
import io
import telegram
import telegram.ext
import requests
import json
import csv
import datetime
import time
import geoip2.database
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Alignment

from userdata_utils import PersistentDataManager
from format_utils import formatDuration, formatStationName, generate_session_text
from session_utils import filterSessionsByProductAndDays, calcSessionsDuration, getSessionDuration
from drova_utils import DrovaClient
from ip_utils import IpTools


bot = telegram.Bot(token=os.environ["TELEGRAM_BOT_TOKEN"])

ip_tool = IpTools()
PDM = PersistentDataManager()

def send_sessions(update, context, edit_message=False, short_mode=False):
    chat_id = update.message.chat_id

    authToken=PDM.getAuthTokensByChatID(chat_id)
    if authToken is None:
        bot.send_message(chat_id=chat_id, text=f"setup me first")
        return
    
    server_id = PDM.getSelectedStation(chatID=chat_id)

    
    limit = PDM.getLimit(chatID=chat_id)
    currentStationName=""
    if server_id:
        currentStationName = PDM.getStationName(chatID=chat_id, stationID=server_id)

    sessions = DrovaClient.getSessions(authToken=authToken, server_id=server_id, limit=limit)
    if sessions:
        message_long_text = " (excluding those shorter than 5 minutes)" if short_mode==True else ""
        message = f"Last {limit} sessions{message_long_text}:\n\n"

        created_on_past = ""

        for i, session in enumerate(reversed(sessions), start=1):
            product_id = session["product_id"]
            game_name = PDM.getProductData(product_id=product_id)
            if game_name == "Unknown game":
                products_data_update(update, context)
                game_name = PDM.getProductData(product_id=product_id)

            serverName=""
            if server_id is None and str(chat_id) in PDM.getStationNames(chatID=chat_id):
                serverName = PDM.getStationName(chatID=chat_id, stationID=session['server_id'])
                if serverName!="":
                    serverName+="\r\n"

            created_on = datetime.datetime.fromtimestamp(
                session["created_on"] / 1000.0
            ).strftime("%Y-%m-%d")

            if not created_on == created_on_past:
                message += f"<strong>{created_on}</strong>:\n"
                created_on_past = created_on
            
            duration = getSessionDuration(session)
            if (not short_mode) or (short_mode and duration > 300):
                message += generate_session_text(limit, i, game_name, serverName, session, ip_tool)

        current_time = datetime.datetime.now().strftime("%H:%M:%S")

        update_text = f"Update {currentStationName} ({current_time})"

        update_callback_data = "update_sessions"
        if short_mode:
            update_callback_data += "_short"

        reply_markup = telegram.InlineKeyboardMarkup(
            [
                [
                    telegram.InlineKeyboardButton(
                        text=update_text, callback_data=update_callback_data
                    )
                ]
            ]
        )
        if edit_message:
            # Modify the original message with the updated session list
            try:
                bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=update.message.message_id,
                    text=message,
                    reply_markup=reply_markup,
                    parse_mode=telegram.ParseMode.HTML,
                )
            except telegram.TelegramError as e:
                pass

        else:
            # Send a new message with the updated session list
            bot.send_message(
                chat_id=chat_id,
                text=message,
                reply_markup=reply_markup,
                parse_mode=telegram.ParseMode.HTML,
            )

    else:
        bot.send_message(chat_id=chat_id, text=f"Error getting sessions")


# Define the callback function for the update sessions button
def update_sessions_callback(update, context):
    query = update.callback_query
    chat_id = query.message.chat_id
    message_id = query.message.message_id
    last_argument = query.data.split("_")[-1]
    short_mode = last_argument == "short"

    if "update_sessions" in query.data:
        # Modify the original message with the updated session list
        send_sessions(query, context, edit_message=True, short_mode=short_mode)
        query.answer()
    else:
        bot.send_message(
            chat_id=chat_id, text="Sorry, I don't understand that command."
        )
        query.answer()


# Set the X-Auth-Token for this chat ID
def set_auth_token(update, context):
    chat_id = update.message.chat_id
    
    if len(context.args)==0:
        bot.send_message(chat_id=chat_id, text=f"add token to command like /token 123-456-789")
        return
    
    token = context.args[0]

    accountInfo = DrovaClient.getAccountInfo(token)
    
    if accountInfo:
        # Store the X-Auth-Token for this chat ID
        PDM.setAuthToken(chat_id,token)
        PDM.setUserID(chat_id,accountInfo['uuid'])
        bot.send_message(chat_id=chat_id, text=f"X-Auth-Token has been set.\r\nПривет, {accountInfo['name']}!")
        products_data_update(update, context)
        updateStationNames(chat_id)
    else:
        bot.send_message(chat_id=chat_id, text=f"Token error, not set.")
    
def removeAuthToken(update, context):
    chat_id = update.message.chat_id
    result=PDM.setAuthToken(chat_id,"-")
    if result:
        bot.send_message(chat_id=chat_id, text=f"Token removed.")


def updateStationNames(chat_id):
    authToken=PDM.getAuthTokensByChatID(chat_id)
    if authToken is None:
        bot.send_message(chat_id=chat_id, text=f"setup me first")
        return
    user_id = PDM.getUserID(chatID=chat_id)
    getServers(authToken,user_id,chat_id)


def handle_start(update, context):
    chat_id = update.message.chat_id

    helpText="""Не добавляйте токены в непроверенных ботов - владелец бота получит информацию из вашего ЛК!
Команды:
/token 123-456-789 - установить токен из qr кода личного кабинета мерчанта
/removeToken - удалить токен пользователя из бота
/current - Краткий список последних сессий по всем станциям
/station [id станции] - выбор станции из списка или ручным вводом её id
/limit N - смена ограничения на вывод сессий
/sessions [short] - просмотр сессий со всех или с выбранной станции
/dumpall - экспорт сессий по серверам
/dumpOnefile - экспорт сессий одним файлом
"""

    bot.send_message(chat_id=chat_id, text=helpText)


def getServers(authToken,user_id,chat_id):
    servers = DrovaClient.getServers(authToken=authToken, user_id=user_id)
    if servers and len(servers) > 0:
        stationNames = {s['uuid']: s['name'] for s in servers}
        PDM.storeStationNames(chat_id, stationNames)
        return servers
    return None




# Define the callback function for the update button
def update_disabled_callback(update, context):
    query = update.callback_query
    chat_id = query.message.chat_id

    if "update_disabled" in query.data:
        # Modify the original message 
        handle_disabled(query, context, edit_message=True)
        query.answer()
    else:
        bot.send_message(
            chat_id=chat_id, text="Sorry, I don't understand that command."
        )
        query.answer()

# Define the callback function for the update button
def update_stationsinfo_callback(update, context):
    query = update.callback_query
    chat_id = query.message.chat_id

    if "update_stationsinfo" in query.data:
        # Modify the original message 
        handle_stationsinfo(query, context, edit_message=True)
        query.answer()
    else:
        bot.send_message(
            chat_id=chat_id, text="Sorry, I don't understand that command."
        )
        query.answer()

def getProductState(product,okState=""):
    info=""
    if not product['enabled']:
        info+=" Not enabled"
    if not product['published']:
        info+=" Not published"
    if not product['available']:
        info+=" Not available"

    if info=="":
        return (False,okState)
    
    return (True,info)

def handle_stationsinfo(update,context, edit_message=False):
    chat_id = update.message.chat_id

    authToken=PDM.getAuthTokensByChatID(chat_id)
    if authToken is None:
        bot.send_message(chat_id=chat_id, text=f"setup me first")
        return
    
    user_id = PDM.getUserID(chatID=chat_id)

    servers=getServers(authToken,user_id,chat_id)
    if servers:
        currentStations=""

        for s in sorted(servers, key=lambda item: item['name']):
            sessions = DrovaClient.getSessions(authToken, s["uuid"], 1)
            if sessions and len(sessions)>0:
                session=sessions[0]

            ips = DrovaClient.getServerIp(authToken, s['uuid'])
            internalIps, externalIps = ip_tool.split_external_ips(ips)

            trial=""
            if  'groups_list' in s and "Free trial volunteers" in s['groups_list']:
                trial=" (Trial)"

            if currentStations!="":
                currentStations+="\r\n\r\n"
            
            currentStations += generate_session_text(s,session,trial,internalIps,externalIps,ip_tool)
            
        current_time = datetime.datetime.now().strftime("%H:%M:%S")
        update_text = f"Update ({current_time})"
        update_callback_data = "update_stationsinfo"

        reply_markup = telegram.InlineKeyboardMarkup(
            [
                [
                    telegram.InlineKeyboardButton(
                        text=update_text, callback_data=update_callback_data
                    )
                ]
            ]
        )
        if edit_message:
            # Modify the original message with the updated session list
            try:
                bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=update.message.message_id,
                    text=currentStations,
                    reply_markup=reply_markup,
                    parse_mode=telegram.ParseMode.HTML,
                )
            except telegram.TelegramError as e:
                pass

        else:
            bot.send_message(chat_id=chat_id, 
                text=currentStations,
                    reply_markup=reply_markup,
                parse_mode=telegram.ParseMode.HTML,
            )
    else:
        bot.send_message(chat_id=chat_id, text=f"Error")  


def handle_disabled(update,context, edit_message=False):
    chat_id = update.message.chat_id

    authToken=PDM.getAuthTokensByChatID(chat_id)
    if authToken is None:
        bot.send_message(chat_id=chat_id, text=f"setup me first")
        return

    user_id = PDM.getUserID(chatID=chat_id)

    servers=getServers(authToken,user_id,chat_id)
    if not servers is None:

        currentProducts=""

        for s in servers:
            products=DrovaClient.getServerProducts(authToken,user_id,s['uuid'])
            if not products is None and len(products)>0:
                currentServerProducts=""
                for product in products:
                    if not product['published'] or not product['enabled'] or not product['available']:
                        _,info=getProductState(product)
                        currentServerProducts+=product['title']+info+"\r\n"
                if currentServerProducts!="":
                    currentProducts+=f"{formatStationName(s,None)}:\r\n{currentServerProducts}"

        if currentProducts=="":
            currentProducts="all products fine"

        current_time = datetime.datetime.now().strftime("%H:%M:%S")
        update_text = f"Update ({current_time})"
        update_callback_data = "update_disabled"

        reply_markup = telegram.InlineKeyboardMarkup(
            [
                [
                    telegram.InlineKeyboardButton(
                        text=update_text, callback_data=update_callback_data
                    )
                ]
            ]
        )
        if edit_message:
            # Modify the original message with the updated session list
            try:
                bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=update.message.message_id,
                    text=currentProducts,
                    reply_markup=reply_markup,
                    parse_mode=telegram.ParseMode.HTML,
                )
            except telegram.TelegramError as e:
                pass

        else:
            bot.send_message(chat_id=chat_id, 
                text=currentProducts,
                    reply_markup=reply_markup,
                parse_mode=telegram.ParseMode.HTML,
            )
    else:
        bot.send_message(chat_id=chat_id, text=f"Error")                        

# Define the callback function for the update button
def update_current_callback(update, context):
    query = update.callback_query
    chat_id = query.message.chat_id

    if "update_current" in query.data:
        # Modify the original message 
        handle_current(query, context, edit_message=True)
        query.answer()
    else:
        bot.send_message(
            chat_id=chat_id, text="Sorry, I don't understand that command."
        )
        query.answer()



# Set up the command handler for the '/current' command
def handle_current(update, context, edit_message=False):
    chat_id = update.message.chat_id

    authToken=PDM.getAuthTokensByChatID(chat_id)
    if authToken is None:
        bot.send_message(chat_id=chat_id, text=f"setup me first")
        return

    user_id = PDM.getUserID(chatID=chat_id)

    servers=getServers(authToken,user_id,chat_id)
    if not servers is None:
        
        currentSessions=""

        for s in sorted(servers, key=lambda item: item['name']):
            sessions = DrovaClient.getSessions(authToken, s["uuid"], 1)  

            if sessions:
                if len(sessions)>0:
                    for session in sessions:
                        game_name = PDM.getProductData(product_id=session["product_id"])
                        if game_name == "Unknown":
                            products_data_update(update, context)
                            game_name = PDM.getProductData(product_id=session["product_id"])       

                        trial=""
                        if session['billing_type']=="trial":
                            trial=" | Trial"

                        created_on=datetime.datetime.fromtimestamp(session["created_on"] / 1000.0   ).strftime("%d.%m %H:%M")
                        clientCityRange=ip_tool.calcRangeByIp(s,session["creator_ip"])
                        if clientCityRange==-1:
                            clientCityRange=""
                        else:
                            clientCityRange=f" {clientCityRange} км |"
                        currentSessions += formatStationName( s,session) +" | "+game_name+ trial +" | "+ip_tool.getCityByIP(session["creator_ip"])+f" |{clientCityRange} "+created_on+" ("+formatDuration(getSessionDuration(session))+")\r\n"
                else:
                    currentSessions += formatStationName(s,None) +" no sessions\r\n"

        if currentSessions!="":
            current_time = datetime.datetime.now().strftime("%H:%M:%S")
            update_text = f"Update ({current_time})"
            update_callback_data = "update_current"

            reply_markup = telegram.InlineKeyboardMarkup(
                [
                    [
                        telegram.InlineKeyboardButton(
                            text=update_text, callback_data=update_callback_data
                        )
                    ]
                ]
            )
            if edit_message:
                # Modify the original message with the updated session list
                try:
                    bot.edit_message_text(
                        chat_id=chat_id,
                        message_id=update.message.message_id,
                        text=currentSessions,
                        reply_markup=reply_markup,
                        parse_mode=telegram.ParseMode.HTML,
                    )
                except telegram.TelegramError as e:
                    pass

            else:
                bot.send_message(chat_id=chat_id, 
                    text=currentSessions,
                        reply_markup=reply_markup,
                    parse_mode=telegram.ParseMode.HTML,
                )
    else:
        bot.send_message(chat_id=chat_id, text=f"Error")




# Set up the command handler for the '/station' command
def handle_station(update, context):
    chat_id = update.message.chat_id
    
    authToken=PDM.getAuthTokensByChatID(chat_id)
    if authToken is None:
        bot.send_message(chat_id=chat_id, text=f"setup me first")
        return

    # If the user provided an ID, update the params with the new station ID
    if len(context.args) > 0:
        station_id = context.args[0]
        PDM.setSelectedStationID(chat_id,station_id)
        # Send a message to the user confirming the update
        bot.send_message(chat_id=chat_id, text=f"Station ID updated to {station_id}.")
    else:
        user_id = PDM.getUserID(chatID=chat_id)

        
        servers = DrovaClient.getServers(authToken, user_id)
        if servers:
            servers.append({'uuid':"-","name":"all"})

            # Create inline keyboard buttons for each available server ID
            keyboard = [
                [
                    telegram.InlineKeyboardButton(
                        text=s["name"], callback_data=f'set_server_id_{s["uuid"]}'
                    )
                    for s in servers
                ]
            ]
            # Send a message to the user with the list of server IDs as inline keyboard buttons
            reply_markup = telegram.InlineKeyboardMarkup(keyboard)
            bot.send_message(
                chat_id=chat_id, text="Select a station:", reply_markup=reply_markup
            )
        else:
            bot.send_message(chat_id=chat_id, text=f"Error while getting servers list")

def handle_limit(update, context):
    chat_id = update.message.chat_id

    authToken=PDM.getAuthTokensByChatID(chat_id)
    if authToken is None:
        bot.send_message(chat_id=chat_id, text=f"setup me first")
        return
    
    if len(context.args) > 0:
        limit = context.args[0]
        PDM.setLimit(chat_id,limit)
        # Send a message to the user confirming the update
        bot.send_message(chat_id=chat_id, text=f"Limit updated to {limit}.")
    else:
        bot.send_message(chat_id=chat_id, text=f"add limit number to command")



def handle_dumpstantionsproducts(update,context):
    chat_id = update.message.chat_id

    authToken=PDM.getAuthTokensByChatID(chat_id)
    if authToken is None:
        bot.send_message(chat_id=chat_id, text=f"setup me first")

    user_id = PDM.getUserID(chatID=chat_id)

    withTime=False
    daysLimit=0
    if(update['message']['text']).lower()=="/dumpstationsproductswithtime" :
        withTime=True
        allsessions={}
    if(update['message']['text']).lower()=="/dumpstationsproductsmonth" :
        withTime=True
        allsessions={}
        daysLimit=30



    servers=getServers(authToken,user_id,chat_id)
    if servers:

        columns={}
        allProducts={}
        firstSessionDate=datetime.datetime.now()
        LastSessionDate=datetime.datetime.strptime("01/01/2020", "%d/%m/%Y")

        for s in servers:
            columns[s['name']]=0

            products=DrovaClient.getServerProducts(authToken,user_id,s['uuid'])
            if not products is None and len(products)>0:
                for product in products:
                    if not product['title'] in allProducts:
                        allProducts[product['title']] ={}
                    allProducts[product['title']][s['uuid']]=product

            if withTime:
                sessions= DrovaClient.getSessions(authToken,s['uuid'])
                if not sessions is None:
                    allsessions[s['uuid']]=sessions
                if daysLimit==0:
                    for session in sessions:
                        sessionStart=datetime.datetime.fromtimestamp(session["created_on"] / 1000.0)
                        if sessionStart>LastSessionDate:
                            LastSessionDate=sessionStart
                        if sessionStart<firstSessionDate:
                            firstSessionDate=sessionStart

        
        if len(allProducts.keys())>0:

            colN=2
           
            wb = Workbook()
            ws = wb.active

            ws.cell(row=1,column=1).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=1,column=1).alignment = Alignment(horizontal='center')

            if  withTime:
                if daysLimit>0:
                    ws.cell(row=1,column=1).value=ws.cell(row=1,column=1).value+f" данные за {daysLimit} дней"
                else:
                    ws.cell(row=1,column=1).value=ws.cell(row=1,column=1).value+f" данные с {firstSessionDate.strftime('%Y-%m-%d')} по {LastSessionDate.strftime('%Y-%m-%d')}"\
                    +"\r\nу разных станций доступно разное количество сессий, дата начала взята с самой старой"
                    ws.cell(row=1,column=1).alignment = Alignment(wrapText=True)

            #for stationName in sorted(columns.keys()):
            for stationID in sorted(PDM.getStationNames(chatID=chat_id).values()):
                ws.cell(row=1,column=colN).value=stationID
                columns[stationID]=colN
                colN+=1


            names=sorted(allProducts.keys())
            for idx, productName in enumerate(names):
                ws.cell(row=idx+2,column=1).value=productName
                for stationID in allProducts[productName].keys():
                    stateError,cellValue=getProductState( allProducts[productName][stationID],"Active")
                    #ws.cell(row=idx+2,column=columns[stationName]).value=state
                    
                    if  withTime:
                        stationSessions=allsessions[stationID]
                        productID=allProducts[productName][stationID]['productId']
                        productSessions=filterSessionsByProductAndDays(stationSessions,productID,daysLimit)
                        cellValue=formatDuration( calcSessionsDuration(productSessions),False)
                        ws.cell(row=idx+2,column=columns[PDM.getStationName(chatID=chat_id, stationID=stationID)]).number_format ="[h]:mm:ss" #'d h:mm:ss'

                    
                    ws.cell(row=idx+2,column=columns[PDM.getStationName(chatID=chat_id, stationID=stationID)]).value=cellValue

                    if stateError:
                        yellow = "00FFFF00"
                        ws.cell(row=idx+2,column=columns[PDM.getStationName(chatID=chat_id, stationID=stationID)]).fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

            # добавляем формулы
            if ws.max_row>1:
                lastColumn=ws.max_column+1
                #lastRow=ws.max_row+1

                for row in ws.rows:
                    rowNum=0
                    formula=""
                    for cell in row:
                        rowNum=cell.row
                        if cell.col_idx==2 and rowNum>1:
                            formula=f"=SUM({cell.coordinate}"
                        elif cell.col_idx>2 and rowNum>1:
                            formula+=f"+{cell.coordinate}"

                    if formula!="":
                        formula+=")"
                    ws.cell(row=rowNum,column=lastColumn).value=formula
                    ws.cell(row=rowNum,column=lastColumn).number_format ="[h]:mm:ss" #'d h:mm:ss'

                ws.cell(row=1,column=lastColumn).value="Всего"
                #ws.cell(row=lastRow,column=lastColumn).value="fff"
                #ws.cell(row=lastRow,column=lastColumn-1).value="Итого"

            # подбираем ширину колонок
            if ws.max_row>1:
                dims = {}
                for row in ws.rows:
                    for cell in row:
                        if cell.value:
                            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                for col, value in dims.items():
                    ws.column_dimensions[col].width = value*1.1
                
                filename=f"productStates{user_id}.xlsx"
                if withTime:
                    filename=f"productStatesWithTime{user_id}.xlsx"
                if daysLimit>0:
                    filename=f"productStatesDays{daysLimit}_{user_id}.xlsx"

                # old sending via tempfile disabled - check!
                #wb.save(filename)
                #bot.send_document(chat_id=chat_id, document=open(filename, "rb"))

                # testing direct sending
                buf = io.BytesIO()
                buf.name = filename
                wb.save(buf)
                buf.seek(0)
                bot.send_document(chat_id=chat_id, document=buf)

        else:
            bot.send_message(chat_id=chat_id, text=f"Error")
    else:
        bot.send_message(chat_id=chat_id, text=f"Error")



def handle_dump(update, context):
    chat_id = update.message.chat_id

    authToken=PDM.getAuthTokensByChatID(chat_id)
    if authToken is None:
        bot.send_message(chat_id=chat_id, text=f"setup me first")
        return
    
    dumpOnefile=False
    if(update['message']['text']).lower()=="/dumponefile":
        dumpOnefile=True


    if len(context.args) == 0:

        user_id = user_id = PDM.getUserID(chatID=chat_id)

        servers = DrovaClient.getServers(authToken, user_id)
        if servers:
            fieldnames = ['Game name','creator_ip','City','RangeKm','ASN','Date','Duration','Start time','Finish time', 'billing_type','status',  'abort_comment', 'client_id','id','uuid',  'server_id', 'merchant_id', 'product_id', 'created_on', 'finished_on', 'score', 'score_reason', 'score_text', 'parent', 'sched_hints']

            if dumpOnefile:
                wb = Workbook()
                ws = wb.active
                #ws.append(fieldnames)

            for s in servers:
                sessions = DrovaClient.getSessions(authToken, s["uuid"])
                if sessions:
                    for item in sessions:
                        product_id = item.get("product_id")

                        creator_ip = item.get("creator_ip")
                        creator_city= ip_tool.getCityByIP(creator_ip,"X")
                        clientCityRange=ip_tool.calcRangeByIp(s,creator_ip)

                        creator_org= ip_tool.getOrgByIP(creator_ip,"X")

                        game_name = PDM.getProductData(product_id=product_id)


                        created_on = datetime.datetime.fromtimestamp(
                            item["created_on"] / 1000.0
                        ).strftime("%Y-%m-%d")
                        start_time = datetime.datetime.fromtimestamp(
                            item["created_on"] / 1000.0
                        ).strftime("%H:%M:%S")
                        finish_time = item["finished_on"]
                        if finish_time:
                            finish_time = datetime.datetime.fromtimestamp(
                                finish_time / 1000.0
                            ).strftime("%H:%M:%S")
                            duration = datetime.timedelta(
                                    seconds=(item["finished_on"] - item["created_on"]) / 1000
                                )
                        else:
                            finish_time = "Now"
                            duration = datetime.timedelta(
                                    seconds=datetime.datetime.utcnow().timestamp() - item["created_on"] / 1000
                                )
                        #duration_str = str(duration).split(".")[0]
                        duration_str =formatDuration( getSessionDuration(item),False)

                        item["Game name"] = game_name
                        item["City"] = creator_city
                        item["ASN"] = creator_org
                        item["Duration"] = duration_str
                        item["Start time"] = start_time
                        item["Finish time"] = finish_time
                        item["Date"] = created_on
                        item["RangeKm"]=clientCityRange


                        if dumpOnefile:
                            item['Station Name']=s['name']
                            item['created_on']= datetime.datetime.fromtimestamp(item["created_on"] / 1000.0   ).strftime("%Y-%m-%d %H:%M:%S")
                            if not item["finished_on"]  is None:
                                item['finished_on']=datetime.datetime.fromtimestamp(item["finished_on"] / 1000.0   ).strftime("%Y-%m-%d %H:%M:%S")

                    csv_file = "sessions-" + s["name"] + ".csv"


                    if not dumpOnefile:
                        # direct send testing
                        try:
                            s = io.StringIO()
                            writer = csv.DictWriter(s, fieldnames=fieldnames)
                            writer.writeheader()
                            writer.writerows(sessions)
                            s.seek(0)
                            buf = io.BytesIO()  
                            buf.write(s.getvalue().encode())
                            buf.seek(0)  
                            buf.name = csv_file
                            bot.send_document(chat_id=chat_id, document=buf)
                        except Exception as e:
                            bot.send_message(chat_id=chat_id, text=f"Error {e} with station {s['name']}")

                        # old sending via tempfile
                        # Write session data to CSV
                        #with open(csv_file, 'w', newline='') as file:
                        #    writer = csv.DictWriter(file, fieldnames=fieldnames)
                        #    writer.writeheader()
                        #    writer.writerows(sessions)
                        #bot.send_document(chat_id=chat_id, document=open(csv_file, "rb"))
                    else:
                        for row in sessions:
                            if ws.max_row==1:
                                ws.append(list(row.keys()))
                            if 'parent' in row:
                                del row['parent']
                            if 'sched_hints' in row:
                                del row['sched_hints']
                            ws.append(list(row.values()))

            if dumpOnefile:
                # подбираем ширину колонок
                if ws.max_row>1:
                    dims = {}
                    for row in ws.rows:
                        for cell in row:
                            if cell.value:
                                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                    for col, value in dims.items():
                        ws.column_dimensions[col].width = value*1.1
                    
                    # old sending via tempfile disabled - check!
                    #wb.save(f"data{user_id}.xlsx")
                    #bot.send_document(chat_id=chat_id, document=open(f"data{user_id}.xlsx", "rb"))

                    # testing direct sending
                    buf = io.BytesIO()
                    buf.name = f"data{user_id}.xlsx"
                    wb.save(buf)
                    buf.seek(0)
                    bot.send_document(chat_id=chat_id, document=buf)



        else:
            bot.send_message(chat_id=chat_id, text=f"Error: {response.status_code}")

    



def products_data_update(update, context):
    chat_id=update.message.chat_id
    products_data_len_old, products_data_len_new = PDM.updateProductsData()
    
    if products_data_len_old != products_data_len_new:
        bot.send_message(
            chat_id=chat_id,
            text=f"Game database has been updated from {products_data_len_old} games to {products_data_len_new}",
        )
    else:
        bot.send_message(chat_id=chat_id, text=f"Game database is up to date or there has been an error while updating game database")


# Define the callback function for the set server ID buttons
def set_server_id_callback(update, context):
    query = update.callback_query
    chat_id = query.message.chat_id
    server_id = query.data.split("_")[-1]

    # Store the server ID for this chat ID
    PDM.setSelectedStationID(chat_id,server_id)

    if server_id=="-":
        message=f"Selected all stations."
    else:
        message=f"Station ID updated to {server_id}."

    bot.answer_callback_query(
        callback_query_id=query.id, text=message
    )


# Set up the command handler
def handle_command(update, context):
    command = update.message.text

    if "/sessions" in command:
        if len(context.args) > 0 and context.args[0] == "short":
            send_sessions(update, context, short_mode=True)
        else:
            send_sessions(update, context)
    elif command.startswith("/token"):
        set_auth_token(update, context)
    else:
        bot.send_message(
            chat_id=update.message.chat_id,
            text="Sorry, I don't understand that command.",
        )


# Set up the message handler
def handle_message(update, context):
    bot.send_message(
        chat_id=update.message.chat_id, text="Sorry, I don't understand that message."
    )


#  Set up the main function to handle updates
def main():

    updater = telegram.ext.Updater(
        token=os.environ["TELEGRAM_BOT_TOKEN"], use_context=True
    )
    dispatcher = updater.dispatcher

    # Add handlers for the '/sessions' command and regular text messages
    command_handler = telegram.ext.CommandHandler("sessions", handle_command)
    dispatcher.add_handler(command_handler)

    message_handler = telegram.ext.MessageHandler(
        telegram.ext.Filters.text & (~telegram.ext.Filters.command), handle_message
    )
    dispatcher.add_handler(message_handler)

    # Add a handler for the '/token' command
    token_handler = telegram.ext.CommandHandler("token", set_auth_token)
    dispatcher.add_handler(token_handler)

    # Add a handler for the '/removeToken' command
    remove_token_handler = telegram.ext.CommandHandler("removeToken", removeAuthToken)
    dispatcher.add_handler(remove_token_handler)

    # Set up the callback query handlers
    update_sessions_handler = telegram.ext.CallbackQueryHandler(
        update_sessions_callback, pattern="^update_sessions"
    )
    dispatcher.add_handler(update_sessions_handler)

    set_server_id_handler = telegram.ext.CallbackQueryHandler(
        set_server_id_callback, pattern="^set_server_id_"
    )
    dispatcher.add_handler(set_server_id_handler)

    # Set up the callback query handlers
    update_current_handler = telegram.ext.CallbackQueryHandler(
        update_current_callback, pattern="^update_current"
    )
    dispatcher.add_handler(update_current_handler)

    # Set up the command handler for the '/current' command
    current_handler = telegram.ext.CommandHandler("current", handle_current)
    dispatcher.add_handler(current_handler)

    # Set up the callback query handlers
    update_disabled_handler = telegram.ext.CallbackQueryHandler(
        update_disabled_callback, pattern="^update_disabled"
    )
    dispatcher.add_handler(update_disabled_handler)

    # Set up the command handler for the '/disabled' command
    disabled_handler = telegram.ext.CommandHandler("disabled", handle_disabled)
    dispatcher.add_handler(disabled_handler)

    # Set up the callback query handlers
    update_stationsinfo_handler = telegram.ext.CallbackQueryHandler(
        update_stationsinfo_callback, pattern="^update_stationsinfo"
    )
    dispatcher.add_handler(update_stationsinfo_handler)

    # Set up the command handler for the '/stationsinfo' command
    stationsinfo_handler = telegram.ext.CommandHandler("stationsInfo", handle_stationsinfo)
    dispatcher.add_handler(stationsinfo_handler)

    # Set up the command handler for the '/start' command
    start_handler = telegram.ext.CommandHandler("start", handle_start)
    dispatcher.add_handler(start_handler)

    # Set up the command handler for the '/station' command
    station_handler = telegram.ext.CommandHandler("station", handle_station)
    dispatcher.add_handler(station_handler)

    # Set up the command handler for the '/limit' command
    limit_handler = telegram.ext.CommandHandler("limit", handle_limit)
    dispatcher.add_handler(limit_handler)

    # Set up the command handler for the '/dumpall' command
    dump_handler = telegram.ext.CommandHandler("dumpall", handle_dump)
    dispatcher.add_handler(dump_handler)

    dump_handler2 = telegram.ext.CommandHandler("dumpOnefile", handle_dump)
    dispatcher.add_handler(dump_handler2)

    dump_stantionsproducts = telegram.ext.CommandHandler("dumpStationsProducts", handle_dumpstantionsproducts)
    dispatcher.add_handler(dump_stantionsproducts)

    dump_stantionsproducts2 = telegram.ext.CommandHandler("dumpStationsProductsWithTime", handle_dumpstantionsproducts)
    dispatcher.add_handler(dump_stantionsproducts2)
    dump_stantionsproducts3 = telegram.ext.CommandHandler("dumpStationsProductsMonth", handle_dumpstantionsproducts)
    dispatcher.add_handler(dump_stantionsproducts3)

    updater.start_polling()
    updater.idle()


if __name__ == "__main__":
    main()
