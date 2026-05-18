from __future__ import annotations

import csv
from datetime import datetime, timedelta
from io import BytesIO, StringIO

import pytest
from openpyxl import load_workbook

from drova_bot.domain.models import Session, Station, StationProduct
from drova_bot.exports.products import ProductExportService
from drova_bot.exports.sessions import SESSION_EXPORT_HEADERS, SessionExportService


@pytest.mark.asyncio
async def test_sessions_xlsx_headers_cells_and_filename(
    ui_sessions: list[Session],
    ui_stations: list[Station],
    ui_catalog: dict[str, str],
    ui_now: datetime,
) -> None:
    export = await SessionExportService().build_sessions_xlsx(
        sessions=ui_sessions,
        stations=ui_stations,
        product_catalog=ui_catalog,
        now=ui_now,
        timezone="Asia/Yekaterinburg",
    )

    workbook = load_workbook(BytesIO(export.payload))
    sheet = workbook["sessions"]

    assert export.filename == "drova-sessions-20260518-120000.xlsx"
    assert [cell.value for cell in sheet[1]] == SESSION_EXPORT_HEADERS
    assert sheet["A2"].value == "Gamma Trial"
    assert sheet["B2"].value == "Space Farm"
    assert sheet["H2"].value == "00:20:00"
    assert sheet["J2"].value is None
    assert sheet["X2"].value is None
    assert sheet["A3"].value == "Alpha Station"
    assert sheet["B3"].value == "Cyber Rally"
    assert sheet["H3"].value == "00:10:00"
    assert sheet["W3"].value == "ok"


@pytest.mark.asyncio
async def test_sessions_csv_per_station_headers_and_sanitized_filename(
    ui_sessions: list[Session],
    ui_stations: list[Station],
    ui_catalog: dict[str, str],
    ui_now: datetime,
) -> None:
    files = await SessionExportService().build_sessions_csv_by_station(
        sessions=ui_sessions,
        stations=ui_stations,
        product_catalog=ui_catalog,
        now=ui_now,
        timezone="Asia/Yekaterinburg",
    )

    assert [file.filename for file in files] == [
        "drova-sessions-alpha-station-20260518-120000.csv",
        "drova-sessions-beta-test-station-20260518-120000.csv",
        "drova-sessions-gamma-trial-20260518-120000.csv",
    ]
    alpha_rows = list(csv.reader(StringIO(files[0].payload.decode("utf-8"))))
    assert alpha_rows[0] == SESSION_EXPORT_HEADERS
    assert alpha_rows[1][0] == "Alpha Station"
    assert alpha_rows[1][1] == "Cyber Rally"
    assert alpha_rows[2][1] == "Desktop Mode"


@pytest.mark.asyncio
async def test_products_xlsx_sorted_matrix_and_problem_fill(
    ui_stations: list[Station],
    ui_products_by_station: dict[str, list[StationProduct]],
    ui_now: datetime,
) -> None:
    export = await ProductExportService().build_products_xlsx(
        stations=ui_stations,
        products_by_station=ui_products_by_station,
        now=ui_now,
    )

    workbook = load_workbook(BytesIO(export.payload))
    sheet = workbook["products"]

    assert export.filename == "drova-products-20260518-120000.xlsx"
    assert [cell.value for cell in sheet[1]] == [
        "Продукт",
        "Alpha Station",
        "Beta Test Station",
        "Gamma Trial",
    ]
    assert [sheet[f"A{row}"].value for row in range(2, 5)] == [
        "Cyber Rally",
        "Desktop Mode",
        "Space Farm",
    ]
    assert sheet["B2"].value == "Active"
    assert sheet["C3"].value == "не опубликован"
    assert sheet["B4"].value == "отключен"
    assert sheet["D4"].value == "недоступен"
    assert sheet["C3"].fill.fill_type == "solid"
    assert sheet["B4"].fill.fill_type == "solid"


@pytest.mark.asyncio
async def test_product_time_xlsx_durations_and_total_formula(
    ui_stations: list[Station],
    ui_sessions: list[Session],
    ui_catalog: dict[str, str],
    ui_now: datetime,
) -> None:
    export = await ProductExportService().build_product_time_xlsx(
        stations=ui_stations,
        sessions=ui_sessions,
        product_catalog=ui_catalog,
        now=ui_now,
    )

    workbook = load_workbook(BytesIO(export.payload), data_only=False)
    sheet = workbook["product-time"]

    assert export.filename == "drova-product-time-20260518-120000.xlsx"
    assert [cell.value for cell in sheet[1]] == [
        "Продукт",
        "Alpha Station",
        "Beta Test Station",
        "Gamma Trial",
        "Всего",
    ]
    assert sheet["A2"].value == "Cyber Rally"
    assert sheet["B2"].value == timedelta(minutes=10)
    assert sheet["E2"].value == "=SUM(B2:D2)"
    assert sheet["E2"].number_format == "[h]:mm:ss"
    assert sheet["A4"].value == "Space Farm"
    assert sheet["D4"].value == timedelta(minutes=20)
