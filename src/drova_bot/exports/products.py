"""Product matrix and product-time XLSX export services."""

from __future__ import annotations

import asyncio
from collections import defaultdict
from collections.abc import Mapping, Sequence
from datetime import datetime
from io import BytesIO
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from drova_bot.domain.formatters import (
    product_problem_flags,
    session_duration_seconds,
    sort_stations,
)
from drova_bot.domain.models import Session, Station, StationProduct
from drova_bot.exports.models import ExportFile
from drova_bot.exports.sessions import XLSX_CONTENT_TYPE

PROBLEM_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")


class ProductExportService:
    async def build_products_xlsx(
        self,
        *,
        stations: Sequence[Station],
        products_by_station: Mapping[str, Sequence[StationProduct]],
        now: datetime,
    ) -> ExportFile:
        return await asyncio.to_thread(
            self._build_products_xlsx,
            stations,
            products_by_station,
            now,
        )

    async def build_product_time_xlsx(
        self,
        *,
        stations: Sequence[Station],
        sessions: Sequence[Session],
        product_catalog: Mapping[str, str],
        now: datetime,
    ) -> ExportFile:
        return await asyncio.to_thread(
            self._build_product_time_xlsx,
            stations,
            sessions,
            product_catalog,
            now,
        )

    @staticmethod
    def products_filename(now: datetime) -> str:
        return f"drova-products-{now.strftime('%Y%m%d-%H%M%S')}.xlsx"

    @staticmethod
    def product_time_filename(now: datetime) -> str:
        return f"drova-product-time-{now.strftime('%Y%m%d-%H%M%S')}.xlsx"

    def _build_products_xlsx(
        self,
        stations: Sequence[Station],
        products_by_station: Mapping[str, Sequence[StationProduct]],
        now: datetime,
    ) -> ExportFile:
        ordered_stations = sort_stations(stations)
        product_titles = sorted(
            {
                product.title
                for station_products in products_by_station.values()
                for product in station_products
            },
            key=str.casefold,
        )
        product_by_station_and_title = {
            (station_id, product.title): product
            for station_id, station_products in products_by_station.items()
            for product in station_products
        }

        workbook = Workbook()
        worksheet = cast(Worksheet, workbook.active)
        worksheet.title = "products"
        worksheet.append(["Продукт", *[station.name for station in ordered_stations]])
        for product_title in product_titles:
            row_index = worksheet.max_row + 1
            worksheet.cell(row=row_index, column=1, value=product_title)
            for column_index, station in enumerate(ordered_stations, start=2):
                product = product_by_station_and_title.get((station.uuid, product_title))
                value = _product_state_cell(product)
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                if value and value != "Active":
                    cell.fill = PROBLEM_FILL

        output = BytesIO()
        workbook.save(output)
        return ExportFile(
            filename=self.products_filename(now),
            content_type=XLSX_CONTENT_TYPE,
            payload=output.getvalue(),
        )

    def _build_product_time_xlsx(
        self,
        stations: Sequence[Station],
        sessions: Sequence[Session],
        product_catalog: Mapping[str, str],
        now: datetime,
    ) -> ExportFile:
        ordered_stations = sort_stations(stations)
        station_column_by_id = {
            station.uuid: index for index, station in enumerate(ordered_stations, start=2)
        }
        product_titles_by_id = {
            session.product_id: product_catalog.get(session.product_id, "Неизвестная игра")
            for session in sessions
        }
        durations: dict[tuple[str, str], int] = defaultdict(int)
        for session in sessions:
            durations[(session.product_id, session.server_id)] += session_duration_seconds(
                session,
                now,
            )

        workbook = Workbook()
        worksheet = cast(Worksheet, workbook.active)
        worksheet.title = "product-time"
        worksheet.append(["Продукт", *[station.name for station in ordered_stations], "Всего"])
        sorted_products = sorted(
            product_titles_by_id.items(),
            key=lambda item: item[1].casefold(),
        )
        for product_id, title in sorted_products:
            row_index = worksheet.max_row + 1
            worksheet.cell(row=row_index, column=1, value=title)
            for station in ordered_stations:
                column_index = station_column_by_id[station.uuid]
                seconds = durations.get((product_id, station.uuid), 0)
                cell = worksheet.cell(row=row_index, column=column_index, value=seconds / 86_400)
                cell.number_format = "[h]:mm:ss"
            first_station_column = 2
            last_station_column = len(ordered_stations) + 1
            total_cell = worksheet.cell(
                row=row_index,
                column=last_station_column + 1,
                value=f"=SUM({worksheet.cell(row_index, first_station_column).coordinate}:"
                f"{worksheet.cell(row_index, last_station_column).coordinate})",
            )
            total_cell.number_format = "[h]:mm:ss"

        output = BytesIO()
        workbook.save(output)
        return ExportFile(
            filename=self.product_time_filename(now),
            content_type=XLSX_CONTENT_TYPE,
            payload=output.getvalue(),
        )


def _product_state_cell(product: StationProduct | None) -> str:
    if product is None:
        return ""
    flags = product_problem_flags(product)
    return ", ".join(flags) if flags else "Active"
