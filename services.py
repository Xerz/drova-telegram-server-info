import io
import json
import datetime
from typing import Any, Dict, List, Optional, Tuple
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

import api
from storage import persistentData, storeStationNames
from utils import format_duration, get_session_duration
from geo_utils import getCityByIP, getOrgByIP, isRfc1918Ip, calcRangeByIp


logger = logging.getLogger(__name__)
def update_products_data() -> Tuple[Dict[str, str], int, int]:
    logger.debug("Updating products data from file and remote API")
    products_data: Dict[str, str] = {}
    try:
        with open("products.json", "r") as f:
            products_data = json.load(f)
    except Exception:
        products_data = {}

    old_len = len(products_data)
    games, status = api.get_products_full()
    if status == 200 and games is not None:
        new_map: Dict[str, str] = {}
        for game in games:
            new_map[game["productId"]] = game["title"]
        products_data = new_map
        with open("products.json", "w") as f:
            f.write(json.dumps(products_data))
    return products_data, old_len, len(products_data)


def get_servers_and_store_names(auth_token: str, user_id: str, chat_id: int) -> Tuple[Optional[List[Dict[str, Any]]], int]:
    logger.debug(f"Fetching servers for user_id={user_id}")
    servers, status = api.get_servers(auth_token, user_id)
    if status == 200 and servers is not None:
        if len(servers) > 0:
            stationNames: Dict[str, str] = {}
            for s in servers:
                stationNames[s['uuid']] = s['name']
            storeStationNames(chat_id, stationNames)
    return servers, status


def format_station_name(station: Dict[str, Any], session: Optional[Dict[str, Any]]) -> str:
    name = station["name"]
    if station["state"] != "LISTEN" and station["state"] != "HANDSHAKE" and station["state"] != "BUSY":
        name = f"<s>{name}</s>"
    if not station.get('published', True):
        name = f"<em>{name}</em>"
    if session is not None:
        if session.get("status") == "ACTIVE" or station.get("state") == "HANDSHAKE":
            name = f"<strong>{name}</strong>"
    return name


def build_sessions_message(auth_token: str, chat_id: int, server_id: Optional[str], limit: int, short_mode: bool, products_data: Dict[str, str]) -> Tuple[str, str, bool, int]:
    logger.debug(f"Building sessions message chat_id={chat_id} server_id={server_id} limit={limit} short={short_mode}")
    merchant_id = persistentData['userIDs'][str(chat_id)]
    data, status = api.get_sessions(auth_token, server_id=server_id, merchant_id=merchant_id, limit=limit)
    if status != 200 or data is None:
        return f"Error: {status}", "", False, status

    sessions = data["sessions"]
    message_long_text = " (excluding those shorter than 5 minutes)" if short_mode else ""
    message = f"Last {limit} sessions{message_long_text}:\n\n"

    created_on_past = ""
    station_names = persistentData.get('stationNames', {}).get(str(chat_id), {})
    unknown_missing = False
    currentStationName = ""
    if server_id is not None:
        currentStationName = station_names.get(server_id, "")

    for i, session in enumerate(reversed(sessions), start=1):
        product_id = session.get("product_id")
        game_name = products_data.get(product_id, "Unknown game")
        if game_name == "Unknown game":
            unknown_missing = True

        serverName = ""
        if server_id is None and str(chat_id) in persistentData['stationNames']:
            serverName = persistentData['stationNames'][str(chat_id)].get(session.get("server_id", ""), "")
            if serverName != "":
                serverName += "\r\n"

        creator_ip = session.get("creator_ip", "N/A")
        creator_city = getCityByIP(creator_ip, "X")
        creator_org = getOrgByIP(creator_ip, "X")

        client_id = session.get("client_id", "")[-6:]

        created_on = datetime.datetime.fromtimestamp(session["created_on"] / 1000.0).strftime("%Y-%m-%d")
        start_time = datetime.datetime.fromtimestamp(session["created_on"] / 1000.0).strftime("%H:%M:%S")
        finish_time = session.get("finished_on", None)

        if finish_time:
            finish_time = datetime.datetime.fromtimestamp(finish_time / 1000.0).strftime("%H:%M:%S")
            duration = datetime.timedelta(seconds=(session["finished_on"] - session["created_on"]) / 1000)
        else:
            finish_time = "Now"
            duration = datetime.timedelta(seconds=datetime.datetime.utcnow().timestamp() - session["created_on"] / 1000)
        duration_str = format_duration(get_session_duration(session))

        score_text = session.get("score_text", None)
        billing_text = session.get("billing_type", None)

        if created_on != created_on_past:
            message += f"<strong>{created_on}</strong>:\n"
            created_on_past = created_on

        if (not short_mode) or (short_mode and duration > datetime.timedelta(minutes=5)):
            if billing_text is not None:
                message += f"{limit - i + 1}. <strong>{game_name}</strong>\n"
                message += serverName
                message += f"<code>{creator_ip}</code> <code>{client_id}</code>\n"
                message += f"{creator_city} {creator_org}\n{start_time}-{finish_time} ({duration_str})\n"
                if score_text is not None:
                    message += f"Feedback: {score_text}\n" if score_text is not None else ""
                message += f"{session.get('billing_type', 'N/A')} {session['status'].lower()}\n\n"

    return message, currentStationName, unknown_missing, 200


def get_product_state(product: Dict[str, Any], ok_state: str = "") -> Tuple[bool, str]:
    info = ""
    if not product.get('enabled', True):
        info += " Not enabled"
    if not product.get('published', True):
        info += " Not published"
    if not product.get('available', True):
        info += " Not available"
    if info == "":
        return False, ok_state
    return True, info


def build_stations_info_message(auth_token: str, user_id: str, chat_id: int) -> Tuple[str, int]:
    logger.debug(f"Building stations info message chat_id={chat_id}")
    servers, status = get_servers_and_store_names(auth_token, user_id, chat_id)
    if status != 200 or servers is None:
        return "Error", status

    currentStations = ""
    for s in sorted(servers, key=lambda item: item['name']):
        ips, st2 = api.get_server_endpoints(auth_token, s['uuid'], limit=1)
        externalIps: List[Dict[str, Any]] = []
        internalIps: List[Dict[str, Any]] = []
        if st2 == 200 and ips is not None and len(ips) > 0:
            for ip in ips:
                if isRfc1918Ip(ip['ip']):
                    internalIps.append(ip)
                else:
                    externalIps.append(ip)

        trial = ""
        if 'groups_list' in s and "Free trial volunteers" in s['groups_list']:
            trial = " (Trial)"

        if currentStations != "":
            currentStations += "\r\n\r\n"
        currentStations += f"{format_station_name(s, None)}{trial}:"
        currentStations += f"\r\n {s['city_name']}"

        if len(externalIps) > 0:
            currentStations += "\r\n Внешние адреса:"
            for ip in sorted(externalIps, key=lambda item: item['ip']):
                city = getCityByIP(ip['ip'], "")
                org = getOrgByIP(ip['ip'], "")
                if len(org) > 0:
                    org = f", {org[0:20]}"
                if city != "":
                    city = f"({city[0:15]}{org})"
                currentStations += f"\r\n <code>{ip['ip']}</code>:{ip['base_port']} {city}"
        if len(internalIps) > 0:
            currentStations += "\r\n Внутренние адреса:"
            for ip in sorted(internalIps, key=lambda item: item['ip']):
                currentStations += f"\r\n <code>{ip['ip']}</code>:{ip['base_port']}"

    return currentStations, 200


def build_disabled_products_message(auth_token: str, user_id: str) -> Tuple[str, int]:
    logger.debug(f"Building disabled products message for user_id={user_id}")
    servers, status = api.get_servers(auth_token, user_id)
    if status != 200 or servers is None:
        return "Error", status

    currentProducts = ""
    for s in servers:
        products, st2 = api.get_server_products(auth_token, user_id, s['uuid'])
        if st2 == 200 and products is not None and len(products) > 0:
            currentServerProducts = ""
            for product in products:
                if not product.get('published', True) or not product.get('enabled', True) or not product.get('available', True):
                    _, info = get_product_state(product)
                    currentServerProducts += product['title'] + info + "\r\n"
            if currentServerProducts != "":
                currentProducts += f"{format_station_name(s, None)}:\r\n{currentServerProducts}"

    if currentProducts == "":
        currentProducts = "all products fine"
    return currentProducts, 200


def build_current_message(auth_token: str, user_id: str, chat_id: int, products_data: Dict[str, str]) -> Tuple[str, int]:
    logger.debug(f"Building current sessions message chat_id={chat_id}")
    servers, status = get_servers_and_store_names(auth_token, user_id, chat_id)
    if status != 200 or servers is None:
        return "Error", status

    currentSessions = ""
    for s in sorted(servers, key=lambda item: item['name']):
        data, st2 = api.get_sessions(auth_token, server_id=s["uuid"], limit=1)
        if st2 == 200 and data is not None:
            sessions = data
            if len(sessions["sessions"]) > 0:
                for session in sessions["sessions"]:
                    game_name = products_data.get(session["product_id"], "Unknown")
                    trial = ""
                    if session.get('billing_type') == "trial":
                        trial = " | Trial"
                    created_on = datetime.datetime.fromtimestamp(session["created_on"] / 1000.0).strftime("%d.%m %H:%M")
                    clientCityRange = calcRangeByIp(s, session["creator_ip"])
                    if clientCityRange == -1:
                        clientCityRange = ""
                    else:
                        clientCityRange = f" {clientCityRange} км |"
                    currentSessions += format_station_name(s, session) + " | " + game_name + trial + " | " + getCityByIP(session["creator_ip"]) + f" |{clientCityRange} " + created_on + " (" + format_duration(get_session_duration(session)) + ")\r\n"
            else:
                currentSessions += format_station_name(s, None) + " no sessions\r\n"
    return currentSessions, 200


def filter_sessions_by_product_and_days(stationSessions: List[Dict[str, Any]], productID: Any, daysLimit: int = 30) -> List[Dict[str, Any]]:
    monthProductSessions: List[Dict[str, Any]] = []
    monthBack = datetime.datetime.now() - datetime.timedelta(days=daysLimit)
    for session in stationSessions:
        if daysLimit > 0:
            if session['product_id'] == productID and session['created_on'] / 1000 > monthBack.timestamp():
                monthProductSessions.append(session)
        else:
            if session['product_id'] == productID:
                monthProductSessions.append(session)
    return monthProductSessions


def calc_sessions_duration(sessions: List[Dict[str, Any]]) -> float:
    duration = 0.0
    for session in sessions:
        duration += get_session_duration(session)
    return duration


def export_sessions(auth_token: str, user_id: str, dump_one_file: bool, products_data: Dict[str, str]) -> Tuple[List[Tuple[str, io.BytesIO]], int]:
    attachments: List[Tuple[str, io.BytesIO]] = []
    logger.debug(f"Export sessions dump_one_file={dump_one_file} user_id={user_id}")
    servers, status = api.get_servers(auth_token, user_id)
    if status != 200 or servers is None:
        return attachments, status

    fieldnames = ['Game name','creator_ip','City','RangeKm','ASN','Date','Duration','Start time','Finish time', 'billing_type','status',  'abort_comment', 'client_id','id','uuid',  'server_id', 'merchant_id', 'product_id', 'created_on', 'finished_on', 'score', 'score_reason', 'score_text', 'parent', 'sched_hints']

    if dump_one_file:
        wb = Workbook()
        ws = wb.active

    for s in servers:
        data, st2 = api.get_sessions(auth_token, server_id=s["uuid"])
        if st2 == 200 and data is not None:
            sessions = data["sessions"]
            for item in sessions:
                product_id = item.get("product_id")
                creator_ip = item.get("creator_ip")
                creator_city = getCityByIP(creator_ip, "X")
                clientCityRange = calcRangeByIp(s, creator_ip)
                creator_org = getOrgByIP(creator_ip, "X")
                game_name = products_data.get(product_id, "Unknown game")

                created_on = datetime.datetime.fromtimestamp(item["created_on"] / 1000.0).strftime("%Y-%m-%d")
                start_time = datetime.datetime.fromtimestamp(item["created_on"] / 1000.0).strftime("%H:%M:%S")
                finish_time = item["finished_on"]
                if finish_time:
                    finish_time = datetime.datetime.fromtimestamp(finish_time / 1000.0).strftime("%H:%M:%S")
                else:
                    finish_time = "Now"
                duration_str = format_duration(get_session_duration(item), False)

                item["Game name"] = game_name
                item["City"] = creator_city
                item["ASN"] = creator_org
                item["Duration"] = duration_str
                item["Start time"] = start_time
                item["Finish time"] = finish_time
                item["Date"] = created_on
                item["RangeKm"] = clientCityRange

            if not dump_one_file:
                import csv
                csv_file = "sessions-" + s["name"] + ".csv"
                try:
                    buf_text = io.StringIO()
                    writer = csv.DictWriter(buf_text, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(sessions)
                    buf_text.seek(0)
                    buf = io.BytesIO()
                    buf.write(buf_text.getvalue().encode())
                    buf.seek(0)
                    buf.name = csv_file
                    attachments.append((csv_file, buf))
                except Exception:
                    continue
            else:
                for row in sessions:
                    if ws.max_row == 1:
                        ws.append(fieldnames)
                    row_data = [row.get(column, "") for column in fieldnames]
                    ws.append(row_data)

    if dump_one_file:
        if ws.max_row > 1:
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value * 1.1

            buf = io.BytesIO()
            filename = f"data{user_id}.xlsx"
            buf.name = filename
            wb.save(buf)
            buf.seek(0)
            attachments.append((filename, buf))
    return attachments, 200


def export_stations_products(auth_token: str, user_id: str, chat_id: int, with_time: bool, days_limit: int) -> Tuple[Optional[Tuple[str, io.BytesIO]], int]:
    logger.debug(f"Export stations products with_time={with_time} days_limit={days_limit} chat_id={chat_id}")
    servers, status = get_servers_and_store_names(auth_token, user_id, chat_id)
    if status != 200 or servers is None:
        return None, status

    columns: Dict[str, int] = {}
    allProducts: Dict[str, Dict[str, Any]] = {}
    firstSessionDate = datetime.datetime.now()
    LastSessionDate = datetime.datetime.strptime("01/01/2020", "%d/%m/%Y")
    allsessions: Dict[str, List[Dict[str, Any]]] = {}

    for s in servers:
        columns[s['name']] = 0

        products, st2 = api.get_server_products(auth_token, user_id, s['uuid'])
        if st2 == 200 and products is not None and len(products) > 0:
            for product in products:
                if product['title'] not in allProducts:
                    allProducts[product['title']] = {}
                allProducts[product['title']][s['uuid']] = product

        if with_time:
            data, st3 = api.get_sessions(auth_token, server_id=s['uuid'])
            if st3 == 200 and data is not None:
                sessions = data["sessions"]
                allsessions[s['uuid']] = sessions
                if days_limit == 0:
                    for session in sessions:
                        sessionStart = datetime.datetime.fromtimestamp(session["created_on"] / 1000.0)
                        if sessionStart > LastSessionDate:
                            LastSessionDate = sessionStart
                        if sessionStart < firstSessionDate:
                            firstSessionDate = sessionStart

    if len(allProducts.keys()) == 0:
        return None, 200

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    if with_time:
        if days_limit > 0:
            ws.cell(row=1, column=1).value = ws.cell(row=1, column=1).value + f" данные за {days_limit} дней"
        else:
            ws.cell(row=1, column=1).value = ws.cell(row=1, column=1).value + f" данные с {firstSessionDate.strftime('%Y-%m-%d')} по {LastSessionDate.strftime('%Y-%m-%d')}" + "\r\nу разных станций доступно разное количество сессий, дата начала взята с самой старой"
            ws.cell(row=1, column=1).alignment = Alignment(wrapText=True)

    # Column headers are station names from persistent map to keep order stable
    colN = 2
    for stationName in sorted(persistentData['stationNames'][str(chat_id)].values()):
        ws.cell(row=1, column=colN).value = stationName
        columns[stationName] = colN
        colN += 1

    names = sorted(allProducts.keys())
    for idx, productName in enumerate(names):
        ws.cell(row=idx + 2, column=1).value = productName
        for stationID in allProducts[productName].keys():  # stationID is uuid
            stateError, cellValue = get_product_state(allProducts[productName][stationID], "Active")

            if with_time:
                stationSessions = allsessions.get(stationID, [])
                productID = allProducts[productName][stationID]['productId']
                productSessions = filter_sessions_by_product_and_days(stationSessions, productID, days_limit)
                cellValue = format_duration(calc_sessions_duration(productSessions), False)
                ws.cell(row=idx + 2, column=columns[persistentData['stationNames'][str(chat_id)][stationID]]).number_format = "[h]:mm:ss"

            ws.cell(row=idx + 2, column=columns[persistentData['stationNames'][str(chat_id)][stationID]]).value = cellValue

            if stateError:
                yellow = "00FFFF00"
                ws.cell(row=idx + 2, column=columns[persistentData['stationNames'][str(chat_id)][stationID]]).fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")

    # Add total formulas
    if ws.max_row > 1:
        lastColumn = ws.max_column + 1
        for row in ws.rows:
            rowNum = 0
            formula = ""
            for cell in row:
                rowNum = cell.row
                if cell.col_idx == 2 and rowNum > 1:
                    formula = f"=SUM({cell.coordinate}"
                elif cell.col_idx > 2 and rowNum > 1:
                    formula += f"+{cell.coordinate}"
            if formula != "":
                formula += ")"
            ws.cell(row=rowNum, column=lastColumn).value = formula
            ws.cell(row=rowNum, column=lastColumn).number_format = "[h]:mm:ss"
        ws.cell(row=1, column=lastColumn).value = "Всего"

    # autosize columns
    if ws.max_row > 1:
        dims: Dict[str, int] = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value * 1.1

    filename = f"productStates{user_id}.xlsx"
    if with_time:
        filename = f"productStatesWithTime{user_id}.xlsx"
    if days_limit > 0:
        filename = f"productStatesDays{days_limit}_{user_id}.xlsx"

    buf = io.BytesIO()
    buf.name = filename
    wb.save(buf)
    buf.seek(0)
    return (filename, buf), 200
